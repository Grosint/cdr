"""
Excel Export Module
Generates multi-sheet Excel files with exact sheet names matching the analytical views
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
from typing import Dict, List
from cdr_analytics import generate_all_analytics
from typing import Optional


async def export_to_excel(session_id: Optional[str] = None, suspect_name: Optional[str] = None) -> str:
    """
    Export all analytical views to multi-sheet Excel file
    Sheet names must match exactly:
    Summary, Corrected, MaxCall, MaxCircleCall, DailyFirstLast,
    MaxDuration, MaxIMEI, DailyIMEIATracking, MaxLocation, DailyFirstLastLocation
    """
    # Generate all analytics
    analytics_data = await generate_all_analytics(session_id, suspect_name)

    # Create exports directory
    exports_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "exports")
    os.makedirs(exports_dir, exist_ok=True)

    # Create filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    name_part = session_id or suspect_name or "cdr"
    filename = os.path.join(exports_dir, f"{name_part}_cdr_analysis_{timestamp}.xlsx")

    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Define sheet order and names
    sheet_names = [
        "Summary",
        "Corrected",
        "MaxCall",
        "MaxCircleCall",
        "DailyFirstLast",
        "MaxDuration",
        "MaxIMEI",
        "DailyIMEIATracking",
        "MaxLocation",
        "DailyFirstLastLocation"
    ]

    # Create and format each sheet
    for sheet_name in sheet_names:
        data = analytics_data.get(sheet_name, {})
        ws = wb.create_sheet(title=sheet_name)

        if sheet_name == "Summary":
            _write_summary_sheet(ws, data)
        elif sheet_name == "Corrected":
            _write_corrected_sheet(ws, data)
        elif sheet_name == "MaxCall":
            _write_max_call_sheet(ws, data)
        elif sheet_name == "MaxCircleCall":
            _write_max_circle_call_sheet(ws, data)
        elif sheet_name == "DailyFirstLast":
            _write_daily_first_last_sheet(ws, data)
        elif sheet_name == "MaxDuration":
            _write_max_duration_sheet(ws, data)
        elif sheet_name == "MaxIMEI":
            _write_max_imei_sheet(ws, data)
        elif sheet_name == "DailyIMEIATracking":
            _write_daily_imei_tracking_sheet(ws, data)
        elif sheet_name == "MaxLocation":
            _write_max_location_sheet(ws, data)
        elif sheet_name == "DailyFirstLastLocation":
            _write_daily_first_last_location_sheet(ws, data)

    # Save workbook
    wb.save(filename)
    return filename


def _write_summary_sheet(ws, data: Dict):
    """Write Summary sheet with KPI summary and tabular breakdown"""
    ws.append(["CDR Analysis Summary"])
    ws.append([])

    # KPI Summary
    ws.append(["Metric", "Value"])
    ws.append(["Total Calls", data.get("total_calls", 0)])
    ws.append(["Incoming Calls", data.get("incoming_count", 0)])
    ws.append(["Outgoing Calls", data.get("outgoing_count", 0)])
    ws.append(["Unique B-Numbers", data.get("unique_b_numbers", 0)])
    ws.append(["Unique IMEIs", data.get("unique_imeis", 0)])
    ws.append(["Unique Locations", data.get("unique_locations", 0)])
    ws.append(["First Activity Date", data.get("first_activity_date", "N/A")])
    ws.append(["Last Activity Date", data.get("last_activity_date", "N/A")])

    _format_header_row(ws, 3)
    _auto_adjust_columns(ws)


def _write_corrected_sheet(ws, data: List[Dict]):
    """Write Corrected sheet with cleaned dataset"""
    if not data or len(data) == 0:
        ws.append(["No corrected records found"])
        _auto_adjust_columns(ws)
        return

    # Headers
    headers = [
        "record_id", "msisdn_a", "msisdn_b", "call_type", "call_date",
        "call_start_time", "call_duration_sec", "imei", "imsi", "cell_id",
        "lac", "operator", "circle", "location_description", "raw_row_reference"
    ]
    ws.append(headers)
    _format_header_row(ws, 1)

    # Data rows
    for record in data:
        ws.append([
            record.get("record_id", ""),
            record.get("msisdn_a", ""),
            record.get("msisdn_b", ""),
            record.get("call_type", ""),
            record.get("call_date", ""),
            record.get("call_start_time", ""),
            record.get("call_duration_sec", 0),
            record.get("imei", ""),
            record.get("imsi", ""),
            record.get("cell_id", ""),
            record.get("lac", ""),
            record.get("operator", ""),
            record.get("circle", ""),
            record.get("location_description", ""),
            record.get("raw_row_reference", "")
        ])

    _auto_adjust_columns(ws)


def _write_max_call_sheet(ws, data: Dict):
    """Write MaxCall sheet"""
    ws.append(["Most Frequently Called Number"])
    ws.append([])
    ws.append(["B-Number", "Total Call Count"])
    ws.append([data.get("b_number", "N/A"), data.get("total_call_count", 0)])

    _format_header_row(ws, 3)
    _auto_adjust_columns(ws)


def _write_max_circle_call_sheet(ws, data: Dict):
    """Write MaxCircleCall sheet"""
    ws.append(["Circle/State with Highest Activity"])
    ws.append([])
    ws.append(["Circle", "Activity Count"])
    ws.append([data.get("circle", "N/A"), data.get("activity_count", 0)])

    _format_header_row(ws, 3)
    _auto_adjust_columns(ws)


def _write_daily_first_last_sheet(ws, data: List[Dict]):
    """Write DailyFirstLast sheet"""
    if not data or len(data) == 0:
        ws.append(["No daily call data found"])
        _auto_adjust_columns(ws)
        return

    headers = [
        "Date", "First Call Time", "First Call B-Number",
        "Last Call Time", "Last Call B-Number"
    ]
    ws.append(headers)
    _format_header_row(ws, 1)

    for item in data:
        ws.append([
            item.get("date", ""),
            item.get("first_call_time", ""),
            item.get("first_call_b_number", ""),
            item.get("last_call_time", ""),
            item.get("last_call_b_number", "")
        ])

    _auto_adjust_columns(ws)


def _write_max_duration_sheet(ws, data: Dict):
    """Write MaxDuration sheet"""
    ws.append(["Longest Duration Call"])
    ws.append([])
    ws.append(["Field", "Value"])
    ws.append(["B-Number", data.get("b_number", "N/A")])
    ws.append(["Duration (seconds)", data.get("duration_seconds", 0)])
    ws.append(["Date", data.get("date", "N/A")])
    ws.append(["Call Start Time", data.get("call_start_time", "N/A")])
    ws.append(["Cell ID", data.get("cell_id", "N/A")])
    ws.append(["Location Description", data.get("location_description", "N/A")])

    _format_header_row(ws, 3)
    _auto_adjust_columns(ws)


def _write_max_imei_sheet(ws, data: Dict):
    """Write MaxIMEI sheet"""
    ws.append(["IMEI Analysis"])
    ws.append([])
    ws.append(["Max IMEI", data.get("max_imei", "N/A")])
    ws.append(["Max IMEI Call Count", data.get("max_imei_call_count", 0)])
    ws.append(["Total IMEIs", data.get("total_imeis", 0)])
    ws.append(["Multi-Device Usage", "Yes" if data.get("multi_device_usage", False) else "No"])
    ws.append([])

    # IMEI Ranking
    ws.append(["IMEI Ranking"])
    ws.append(["IMEI", "Call Count"])
    _format_header_row(ws, 7)

    for item in data.get("imei_ranking", []):
        ws.append([item.get("imei", ""), item.get("call_count", 0)])

    _auto_adjust_columns(ws)


def _write_daily_imei_tracking_sheet(ws, data: List[Dict]):
    """Write DailyIMEIATracking sheet"""
    if not data or len(data) == 0:
        ws.append(["No daily IMEI tracking data found"])
        _auto_adjust_columns(ws)
        return

    ws.append(["Date", "IMEI", "Call Count"])
    _format_header_row(ws, 1)

    for item in data:
        date_str = item.get("date", "")
        imeis = item.get("imeis", [])

        if not imeis:
            ws.append([date_str, "N/A", 0])
        else:
            for imei_data in imeis:
                ws.append([
                    date_str,
                    imei_data.get("imei", ""),
                    imei_data.get("call_count", 0)
                ])

    _auto_adjust_columns(ws)


def _write_max_location_sheet(ws, data: Dict):
    """Write MaxLocation sheet"""
    ws.append(["Most Frequently Used Location"])
    ws.append([])
    ws.append(["Cell ID", "Usage Count"])
    ws.append([data.get("cell_id", "N/A"), data.get("usage_count", 0)])

    _format_header_row(ws, 3)
    _auto_adjust_columns(ws)


def _write_daily_first_last_location_sheet(ws, data: List[Dict]):
    """Write DailyFirstLastLocation sheet"""
    if not data or len(data) == 0:
        ws.append(["No daily location data found"])
        _auto_adjust_columns(ws)
        return

    headers = [
        "Date", "First Location Cell ID", "First Location Time",
        "First Location Description", "Last Location Cell ID",
        "Last Location Time", "Last Location Description"
    ]
    ws.append(headers)
    _format_header_row(ws, 1)

    for item in data:
        first_loc = item.get("first_location", {})
        last_loc = item.get("last_location", {})

        ws.append([
            item.get("date", ""),
            first_loc.get("cell_id", ""),
            first_loc.get("time", ""),
            first_loc.get("location_description", ""),
            last_loc.get("cell_id", ""),
            last_loc.get("time", ""),
            last_loc.get("location_description", "")
        ])

    _auto_adjust_columns(ws)


def _format_header_row(ws, row_num: int):
    """Format header row with bold font and background color"""
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for cell in ws[row_num]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border


def _auto_adjust_columns(ws):
    """Auto-adjust column widths"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
